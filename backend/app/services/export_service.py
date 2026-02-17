"""PDF and Excel export generation."""

import io
from typing import Optional


class ExportService:
    """Generates PDF and Excel exports from query results."""

    async def export_pdf(
        self,
        title: str,
        insight: str,
        columns: list[str],
        rows: list[list],
        chart_image: Optional[bytes] = None,
    ) -> bytes:
        """Generate a PDF report."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 12))

        # Insight
        if insight:
            elements.append(Paragraph(insight, styles['Normal']))
            elements.append(Spacer(1, 12))

        # Data table
        if columns and rows:
            table_data = [columns] + [list(row) for row in rows[:100]]
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366F1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(table)

        doc.build(elements)
        return buffer.getvalue()

    async def export_excel(
        self,
        title: str,
        insight: str,
        columns: list[str],
        rows: list[list],
    ) -> bytes:
        """Generate an Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["DataMind Export"])
        ws_summary.append(["Title:", title])
        ws_summary.append([])
        ws_summary.append(["Insight:"])
        ws_summary.append([insight or ""])

        # Data sheet
        ws_data = wb.create_sheet("Data")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws_data.append(columns)
        for cell in ws_data[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws_data.append(list(row))

        # Auto-width
        for col in ws_data.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws_data.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
